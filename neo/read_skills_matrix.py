'''
This files is the source code containing classes to read matrices from spreadsheet

Author : R. Thomas
Place : U. of Sheffield
Year: 2023-24
Licence: GPLv3
Pylint: 9.78/10
'''
##Standard library imports
import os ##for the test part
import io ##for the test part
import unittest ##for the test part
import unittest.mock  ##for the test part

###Third party imports
import openpyxl
import numpy

###local imports
from . import hardcoded


class GeneralMatrix:
    '''
    This class is a basic class that gets
    properties of a spreadsheet table
    Other classes are inheriting from this
    one
    '''
    def __init__(self, workbook, sheet_name, first_row=1):
        '''
        Class constructor
        '''
        self.name = sheet_name
        self.cols = 0
        self.rows = 0
        self.sheet = workbook[self.name]
        self.first_row = first_row

    def get_table_shape(self):
        '''
        This method gets the shape of the matrix (rows x columns)

        Parameters
        ----------
        skipped_lines   :   list
                            of int: Lines number to be skipped 
                            default is empty list [no skipped lines]

        Returns
        -------
        None


        Attributes
        ----------
        row     int
                number of rows
        cols    int
                number of columns
        '''
        ###get the first line
        line_to_check = self.first_row

        ###look at the line to check
        #columns
        all_letter = []
        for cell in self.sheet[line_to_check]:
            if cell.value is not None:
                all_letter.append(openpyxl.utils.get_column_letter(cell.column))
        self.cols = len(all_letter)

        #rows. Look at all the columns and take the highest row number with
        #content
        max_row = 0
        for l in all_letter:
            for cell in self.sheet[l]:
                if cell.value is not None and cell.row >= self.first_row\
                        and cell.row > max_row:
                    max_row = cell.row+1 ###we had +1 because python starts
                                         ###at 0 while the spreadhsset starts at 1

        self.rows = max_row - self.first_row

    def get_table(self):
        '''
        This method extracts the complete table. It uses the get_table_shape
        method to find the limits

        Parameters
        ----------
        None

        Return
        ------
        None

        Attributes
        ----------
        table   : list of list of columns
        '''
        ###get the limits
        self.get_table_shape()

        ###get the first line to check
        line_to_check = self.first_row

        ###extract the columns
        columns = []
        for cell in self.sheet[line_to_check]:
            if cell.column <= self.cols:
                if cell.value not in ['Example', None, 'Faculty', '-'] and\
                        'RSE' not in cell.value:
                    columns.append(openpyxl.utils.get_column_letter(cell.column))

        ##and then we extract the content of the columns
        self.table = {}
        for col in columns:
            line = []
            for cell in self.sheet[col]:
                if  cell.row >= self.first_row and\
                    cell.row <= self.rows+self.first_row and\
                    cell.value is not None:
                    line.append(cell.value)
            self.table[col] = line

    def get_rses(self):
        '''
        This method gets the name of all RSEs in the subskill matrix

        Parameter
        ---------
        None

        Return
        ------
        rses    dict
                column letter <==> RSE name
        '''
        ##if the table was not extracted we extract it
        if not hasattr(self, 'table'):
            self.get_table()

        ##then go through the table and get the top of each line
        rses = {}
        for col in self.table.items():
            if col[1][0] != self.name: ###<---remove the skill column
                rses[col[0]] = self.table[col[0]][0]

        return rses

    def get_skills_list(self, column_skills = 'A'):
        '''
        This method gets the name of all the skills listed in the subskill matrix
        We assume that the list starts on cell A-3 and goes on the 1st column

        Parameter
        ---------
        column_skills   :   str
                            column letter with skills.
                            Defaut is 'A'

        Return
        ------
        skills      list
                    list of skills for that matrix
        '''
        skills = []
        for cell in self.sheet[column_skills]:
            if cell.row >= self.first_row and\
               cell.row <= self.rows+self.first_row and\
               cell.value is not None and 'RSE' not in cell.value:
                skills.append(cell.value.replace('*', ''))

        return skills[1:]


    def get_skills_per_person(self, rsecolumn, rsename, skills):
        '''
        THis method gets the skills and proficiency
        for all skills in the matrix for the rse
        given in argument

        Parameter
        ---------
        rsecolumn         :str
                           column name of the rse
        skills            : list
                            of skills in the matrix

        return
        ------
        rse_skills    :   nested dict
                          skills : {'prof':X, 'usage':Y/N}
        '''
        skills_values = []
        for cell in self.sheet[rsecolumn]:
            if cell.value is not None and cell.value != rsename:
                skills_values.append(cell.value)

        ##depending on the type of skills matrix we extract
        ##differently.
        rse_skills = {}
        ##if for each skills we have a single line
        if len(skills) == len(skills_values):
            for index,s in enumerate(skills):
                proficiency = 0
                training = 0
                if skills_values[index] == 'Yes':
                    proficiency = 1

                if skills_values[index] in ['No, but interested', 'No but interested']:
                    training = 1
                rse_skills[s] = {'Proficiency': proficiency, 'Training': training}

        ###if for each skills we have two lines
        elif len(skills) == len(skills_values)//2:
            for index,s in enumerate(skills):
                rse_skills[s] = {'Proficiency': skills_values[index*2] ,\
                                 'usage': skills_values[index*2+1]}

        return rse_skills

    def get_nperson_per_skill(self, rses, skills):
        '''
        Get the number of person per skill
        We add a +1 to the skill score
        if the person as a number higher than 1
        or a Yes
        
        Parameter
        ---------
        skillname   :   str
                        skill for which we want the score

        Return
        ------
        skill_score :   int
                        score for the skill
        '''
        ###loop over all rses
        all_rse_skills = []
        for letter in rses:
            if rses[letter] != self.name: ##<---we remove the column that has the skills
                if self.get_skills_per_person(letter, rses[letter], skills):
                    all_rse_skills.append(self.get_skills_per_person(letter, rses[letter],
                                          skills))
                else:
                    print(hardcoded.BOLD+\
                          f'No data (or data incomplete) for {rses[letter]}.')
        #prepare skills
        skills_prof = dict(zip(skills, numpy.zeros(len(skills))))
        skills_npeople = dict(zip(skills, numpy.zeros(len(skills))))

        for s in skills:
            for e in all_rse_skills:
                skills_prof[s] += e[s]['Proficiency']
                if e[s]['Proficiency'] > 0:
                    skills_npeople[s] += 1

        return skills_prof, skills_npeople


class TestreadingMatrix(unittest.TestCase):
    '''
    Class tat tests the reading of a matrix and the method
    associated to the class
    '''
    ##get the fake matrix file
    dir_path = os.path.dirname(os.path.realpath(__file__))
    Fake_matrix = os.path.join(dir_path, 'tests_files/Fake_SkillsInterestMatrix.xlsx')

    def test_simple_matrix_instance(self):
        '''
        Test if the attributes are setup correctly
        '''
        ###open the Fake file
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HighLevel')

        ##tests
        self.assertEqual(matrix.name, 'HighLevel')
        self.assertEqual(matrix.first_row, 1)
        self.assertEqual(matrix.rows, 0)
        self.assertEqual(matrix.cols, 0)

        ###open the Fake file
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HighLevel', first_row=1)
        self.assertEqual(matrix.first_row, 1)


    def test_simple_matrix_instance_with_wrong_matrix_name(self):
        '''
        #Test if the submatrix name is wrong that it gives an error
        '''
        ###open the Fake file
        opened = openpyxl.load_workbook(self.Fake_matrix)

        ##tests
        self.assertRaises(KeyError, lambda: GeneralMatrix(opened, 'low Level'))

    def test_rows_and_columns_count(self):
        '''
        #Test if the matrix has the right amount of columns and rows
        '''
        ###open the Fake file
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HighLevel', first_row=13)
        matrix.get_table_shape()

        ##tests
        self.assertEqual(matrix.name, 'HighLevel')
        self.assertEqual(matrix.rows, 11)
        self.assertEqual(matrix.cols, 15)

        ###open the Fake file
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HL1', first_row=2)
        matrix.get_table_shape()

        ##tests
        self.assertEqual(matrix.name, 'HL1')
        self.assertEqual(matrix.cols, 8)
        self.assertEqual(matrix.rows, 18)

        ##other spreadsheet
        matrix = GeneralMatrix(opened, 'HL2', first_row=2)
        matrix.get_table_shape()

        ##tests
        self.assertEqual(matrix.name, 'HL2')
        self.assertEqual(matrix.cols, 14)
        self.assertEqual(matrix.rows, 18)

        ##other spreadsheet
        matrix = GeneralMatrix(opened, 'HL3', first_row=2)
        matrix.get_table_shape()

        ##tests
        self.assertEqual(matrix.name, 'HL3')
        self.assertEqual(matrix.cols, 5)
        self.assertEqual(matrix.rows, 10)

        ##other spreadsheet
        matrix = GeneralMatrix(opened, 'Topics', first_row=5)
        matrix.get_table_shape()

        ##tests
        self.assertEqual(matrix.name, 'Topics')
        self.assertEqual(matrix.cols, 15)
        self.assertEqual(matrix.rows, 13)

    def test_table_content_normal_matrix(self):
        '''
        #Test that the extracted table has the right content
        '''
        ###open the Fake file
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HL1', first_row=2)
        matrix.get_table()

        ##tests
        self.assertEqual(matrix.table['A'][0], 'HL1')
        self.assertRaises(KeyError, lambda: matrix.table['B'])
        self.assertEqual(matrix.table['C'][0], 'Rogers')
        self.assertEqual(matrix.table['D'][2], 'Yes')
        self.assertEqual(matrix.table['F'][-1], 'Yes')
        self.assertRaises(KeyError, lambda: matrix.table['E'])


    def test_table_content_highlevel_matrix(self):
        '''
        #Test that the extracted table has the right content
        '''
        ###open the Fake file
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HighLevel', first_row=13)
        matrix.get_table()

        ##tests
        self.assertEqual(matrix.table['A'][0], 'High Level')
        self.assertRaises(KeyError, lambda: matrix.table['B'])
        self.assertEqual(matrix.table['C'][0], 'Rogers')
        self.assertEqual(matrix.table['D'][3], 'Yes')
        self.assertEqual(matrix.table['D'][0], 'Loki')
        self.assertEqual(matrix.table['F'][-1], 'No but interested')
        self.assertRaises(KeyError, lambda: matrix.table['P'])

    def test_table_content_topic_sheet(self):
        '''
        Test the content of the table for the Topics sheet
        '''
        ##other spreadsheet
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'Topics', first_row=5)
        matrix.get_table()

        ##tests
        self.assertRaises(KeyError, lambda: matrix.table['A'])
        self.assertEqual(matrix.table['B'][0], 'Department')
        self.assertEqual(matrix.table['D'][3], 'No')
        self.assertEqual(matrix.table['D'][0], 'Loki')
        self.assertEqual(matrix.table['F'][-1], 'Yes')
        self.assertRaises(KeyError, lambda: matrix.table['L'])


    def test_get_rse_list(self):
        '''
        Tests that the list of RSE is correct
        '''
        ##open spreadsheet
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'Topics', first_row=5)
        matrix.get_table()
        rses = matrix.get_rses()

        ##test
        self.assertEqual(rses['C'], 'Rogers')
        self.assertEqual(rses['E'], 'Tony')
        self.assertEqual(rses['O'], 'Odin')

        ##other spreadsheet
        matrix = GeneralMatrix(opened, 'HighLevel', first_row=13)
        matrix.get_table()
        rses = matrix.get_rses()

        ##test
        self.assertEqual(rses['C'], 'Rogers')
        self.assertEqual(rses['E'], 'Tony')
        self.assertRaises(KeyError, lambda: rses['P'])

        ##other spreadsheet
        matrix = GeneralMatrix(opened, 'HL1', first_row=2)
        matrix.get_table()
        rses = matrix.get_rses()

        ##test
        self.assertEqual(rses['C'], 'Rogers')
        self.assertEqual(rses['H'], 'Thor')
        self.assertRaises(KeyError, lambda: rses['E'])


    def test_get_skills(self):
        '''
        Test the get_skill method

        '''
        ##open spreadsheet
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'Topics', first_row=5)
        matrix.get_table()
        skills = matrix.get_skills_list(column_skills='B')

        self.assertEqual(skills[0], 'dept1')
        self.assertEqual(skills[-1], 'dept12')
        self.assertEqual(skills[1], 'dept2')

        ##other spreadsheet
        matrix = GeneralMatrix(opened, 'HL1', first_row=2)
        matrix.get_table()
        skills = matrix.get_skills_list()

        ##test
        self.assertEqual(skills[0], 'HL1_1')
        self.assertEqual(skills[-1], 'HL1_9')
        self.assertEqual(skills[1], 'HL1_2')

        ##other spreadsheet
        matrix = GeneralMatrix(opened, 'HighLevel', first_row=13)
        matrix.get_table()
        skills = matrix.get_skills_list()

        ##test
        self.assertEqual(skills[0], 'HL1')
        self.assertEqual(skills[-1], 'HL10')
        self.assertEqual(skills[1], 'HL2')

    def test_get_skills_per_person_submatrix(self):
        '''
        Tests that we can get the right skills for a single person
        '''
        ##open spreadsheet
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HL1', first_row=2)
        matrix.get_table()
        rses = matrix.get_rses()
        skills = matrix.get_skills_list()

        item = list(rses.items())[0]
        rse_skills = matrix.get_skills_per_person(item[0], item[1], skills)

        ##tests
        self.assertEqual(item[1], 'Rogers')

        self.assertEqual(rse_skills['HL1_1']['Proficiency'], 3)
        self.assertEqual(rse_skills['HL1_2']['Proficiency'], 3)
        self.assertEqual(rse_skills['HL1_3']['Proficiency'], 3)
        self.assertEqual(rse_skills['HL1_4']['Proficiency'], 2)
        self.assertEqual(rse_skills['HL1_5']['Proficiency'], 1)
        self.assertEqual(rse_skills['HL1_6']['Proficiency'], 0)
        self.assertEqual(rse_skills['HL1_7']['Proficiency'], 0)
        self.assertEqual(rse_skills['HL1_8']['Proficiency'], 1)
        self.assertEqual(rse_skills['HL1_9']['Proficiency'], 2)

        self.assertEqual(rse_skills['HL1_1']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_2']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_3']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_4']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_5']['usage'], 'No')
        self.assertEqual(rse_skills['HL1_6']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_7']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_8']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_9']['usage'], 'No')

    def test_get_skills_per_person_highlevel(self):
        '''
        Tests that we can get the right skills for a single person
        on the high level matrix
        '''
        ##open spreadsheet
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HighLevel', first_row=13)
        matrix.get_table()
        rses = matrix.get_rses()
        skills = matrix.get_skills_list()

        item = list(rses.items())[4]
        rse_skills = matrix.get_skills_per_person(item[0], item[1], skills)

        ##tests
        self.assertEqual(item[1], 'Fury')

        self.assertEqual(rse_skills['HL1']['Proficiency'], 1)
        self.assertEqual(rse_skills['HL2']['Proficiency'], 1)
        self.assertEqual(rse_skills['HL3']['Proficiency'], 0)
        self.assertEqual(rse_skills['HL4']['Proficiency'], 0)

        self.assertEqual(rse_skills['HL1']['Training'], 0)
        self.assertEqual(rse_skills['HL2']['Training'], 0)
        self.assertEqual(rse_skills['HL3']['Training'], 0)
        self.assertEqual(rse_skills['HL4']['Training'], 1)

    def test_get_skills_per_person_topics(self):
        '''
        Tests that we can get the right skills for a single person
        on the high level matrix
        '''
        ##open spreadsheet
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'Topics', first_row=5)
        matrix.get_table()
        rses = matrix.get_rses()
        skills = matrix.get_skills_list(column_skills='B')

        ##get the skills
        item = list(rses.items())[3]
        rse_skills = matrix.get_skills_per_person(item[0], item[1], skills)

        ##tests
        self.assertEqual(item[1], 'Tony')

        self.assertEqual(rse_skills['dept1']['Proficiency'], 1)
        self.assertEqual(rse_skills['dept5']['Proficiency'], 1)
        self.assertEqual(rse_skills['dept10']['Proficiency'], 1)
        self.assertEqual(rse_skills['dept12']['Proficiency'], 0)

        self.assertEqual(rse_skills['dept1']['Training'], 0)
        self.assertEqual(rse_skills['dept5']['Training'], 0)
        self.assertEqual(rse_skills['dept10']['Training'], 0)
        self.assertEqual(rse_skills['dept12']['Training'], 0)


    @unittest.mock.patch('sys.stdout', new_callable=io.StringIO)
    def displayfinal(self, param, exp, mock_stdout):
        '''
        Function that actually run the test
        '''
        matrix = param[-1]
        skill_prof, skills_npeople = matrix.get_nperson_per_skill(param[0], param[1])
        del skill_prof, skills_npeople
        self.assertEqual(mock_stdout.getvalue(), exp)


    def test_get_nperson_per_skills_submatrix(self):
        '''
        Tests that we can get the right skills for a single person
        '''
        ##open spreadsheet
        opened = openpyxl.load_workbook(self.Fake_matrix)
        matrix = GeneralMatrix(opened, 'HL4', first_row=2)
        matrix.get_table()
        rses = matrix.get_rses()
        skills = matrix.get_skills_list()

        printout ='\033[1m'+\
                  '[Neo: Matrix Generation:]\033[0m : No data (or data incomplete) for Thor.\n'
        self.displayfinal([rses, skills, matrix], printout)
