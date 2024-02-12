'''
This is the file with all the meta functions used directly
from command line argument

Author: R. Thomas
Year: 2023-24
Place: U. of Sheffield
Licence: GPLv3
Pylint: 8.76
'''
###Standard library imports
import os ###for testing only
import io
import unittest

##Third party imports
import openpyxl
import numpy

##local imports
from . import read_skills_matrix
from . import plot_matrix
from . import hardcoded

def training_matrix(matrixfile, plot=True):
    '''
    This function is called when training is given as argument
    it plots the training need program.

    Parameters
    ----------
    matrixfile  :   str
                    file with skills matrix

    plot        :   bool
                    If you want to display the plot
    Return
    ------
    needs       : dict
                  skills : Training needs (number)
    '''
    ###open it
    opened = openpyxl.load_workbook(matrixfile)

    ###Training information are in the High level matrix
    if 'HighLevel' in opened.sheetnames:

        sheetname = 'HighLevel'

        ###load the matrix
        matrix = read_skills_matrix.GeneralMatrix(opened, sheetname, first_row=13)
        matrix.get_table()

        ##get the list of topics
        skills = matrix.get_skills_list()

        ##and the list of rses
        rses = matrix.get_rses()

        ###extract the high level matrix
        rse_training_need = {}
        for rse in rses.items():
            global_skills = matrix.get_skills_per_person(rse[0], rse[1], skills)
            rse_training_need[rse[1]] = global_skills

        ##prepare the training need dictionary
        needs = dict(zip(skills, 0*numpy.ones(len(skills))))

        ##fill it
        for n in needs.items():
            topic = n[0]
            for rse in rse_training_need.items():
                if rse[1]: ##<---make sure the dictionary exists
                    needs[topic] += rse_training_need[rse[0]][topic]['Training']

        ##make the plot
        if plot:
            plot_matrix.make_training(needs)

    else:
        print(hardcoded.BOLD + 'High Level matrix not found...exit..' )
        needs = {} ##<---We have to return a dictionary anyway

    return needs


def single_matrix(matrixfile, name, first_row, skill_column, plot=True):
    '''
    This meta function extracts the matrix of a single topic.
    If the plot function is used, the two plots are given:
        - N people vs skills
        - Total proficiency vs skills

    Parameters
    ----------
    matrixfile  :   str
                    file name
    name        :   str
                    name of a matrix in the file
    first_row   :  int
                   first spreadsheet row where the table starts (header line)
    skill_column:   str
                    letter of the column where skills are

    Return
    ------
    matrix      :   dict
                    with matrix data
    '''
    ###prepare a dictionary for the matrix
    skills_prof = {}
    skills_Npeople = {}

    ###open it
    opened = openpyxl.load_workbook(matrixfile)

    ###check that the name is in the spreadsheet
    if name in opened.sheetnames:

        ###load the matrix
        matrix = read_skills_matrix.GeneralMatrix(opened, name, first_row=first_row)
        matrix.get_table()

        ##get the list of topics
        skills = matrix.get_skills_list(skill_column)

        ##and the list of rses
        rses = matrix.get_rses()

        ###get skills per person
        skills_prof, skills_Npeople = matrix.get_nperson_per_skill(rses, skills)
        #print(skills_Npeople)

        ##plot
        if plot:
            if name != 'HighLevel':
                plot_matrix.make_matrix_spiderplot(name, skills_prof,
                                            title=f'RSE {name} matrix\n[Total proficiency/Skill]')

            plot_matrix.make_matrix_barplot(name, skills_Npeople,
                                            title=f'RSE {name} matrix\n[Npeople/Skill]')

    else:
        print(hardcoded.BOLD + f'Matrix {name} not found...exit...' )

    return skills_prof, skills_Npeople


def single_rse(matrixfile, name, first_row, skill_column, rse, plot=True):
    '''
    This meta function extracts the matrix of a single topic and the data of
    a single RSE.
    If the plot function is used a proficiency spider plot is displayed

    Parameters
    ----------
    matrixfile  :   str
                    file name
    name        :   str
                    name of a matrix in the file
    first_row   :  int
                   first spreadsheet row where the table starts (header line)
    skill_column:   str
                    letter of the column where skills are
    rse         :   str
                    name of the rse

    Return
    ------
    matrix      :   dict
                    with matrix data
    '''
    ###prepare a dictionary for the matrix
    rse_skills = {}

    ###open it
    opened = openpyxl.load_workbook(matrixfile)

    ###check that the name is in the spreadsheet
    if name in opened.sheetnames:

        ###load the matrix
        matrix = read_skills_matrix.GeneralMatrix(opened, name, first_row=first_row)
        matrix.get_table()

        ##get the list of topics
        skills = matrix.get_skills_list(skill_column)

        ##and the list of rses
        rses = matrix.get_rses()

        ##check that the RSE name given is in the list
        for rsename in rses.items():
            if rsename[1] == rse:
                rse_skills = matrix.get_skills_per_person(rsename[0], rsename[1], skills)

        ###if empty
        if not rse_skills:
            print(hardcoded.BOLD + f'The {name} matrix for {rse} is empty...exit...')

        else:
            ##plot
            if plot:
                ###reorganise for plot
                rse_skill = {}
                for s in rse_skills:
                    rse_skill[s] = rse_skills[s]['Proficiency']

                ###let's plot
                plot_matrix.make_rse_indiv(rse_skill, rse,
                                           title=f'{name} matrix for {rse}\n[Proficiency/Skill]')

    else:
        print(hardcoded.BOLD + f'Matrix {name} not found...exit...')

    return rse_skills




class Testmetafunctions(unittest.TestCase):
    '''
    Class that tests the meta functions
    We never display any plot during testing...--> plot=False everwhere
    '''
    ##get the fake matrix file
    dir_path = os.path.dirname(os.path.realpath(__file__))
    Fake_matrix = os.path.join(dir_path, 'tests_files/Fake_SkillsInterestMatrix.xlsx')
    Fake_matrix2 = os.path.join(dir_path, 'tests_files/Fake_SkillsInterestMatrix2.xlsx')



    def test_training_matrix_topics(self):
        '''
        Test that the training matrix data are correct
        '''

        ###get the needs dictionary
        needs = training_matrix(self.Fake_matrix, plot=False)

        ###check that the topics are correct:
        all_topics = list(needs.keys())
        self.assertEqual(all_topics[0], 'HL1')
        self.assertEqual(all_topics[1], 'HL2')
        self.assertEqual(all_topics[2], 'HL3')
        self.assertEqual(all_topics[3], 'HL4')
        self.assertEqual(all_topics[4], 'HL5')
        self.assertEqual(all_topics[5], 'HL6')
        self.assertEqual(all_topics[6], 'HL7')
        self.assertEqual(all_topics[7], 'HL8')
        self.assertEqual(all_topics[8], 'HL9')
        self.assertEqual(all_topics[9], 'HL10')

    def test_training_matrix_values(self):
        '''
        Test that the training matrix data are correct
        '''

        ###get the needs dictionary
        needs = training_matrix(self.Fake_matrix, plot=False)

        ###check that the topics are correct:
        self.assertEqual(needs['HL1'], 5)
        self.assertEqual(needs['HL2'], 5)
        self.assertEqual(needs['HL3'], 4)
        self.assertEqual(needs['HL4'], 8)
        self.assertEqual(needs['HL5'], 3)
        self.assertEqual(needs['HL6'], 1)
        self.assertEqual(needs['HL7'], 5)
        self.assertEqual(needs['HL8'], 4)
        self.assertEqual(needs['HL9'], 3)
        self.assertEqual(needs['HL10'], 7)


    @unittest.mock.patch('sys.stdout', new_callable=io.StringIO)
    def displayfinal_nohighlevel(self, param, exp, mock_stdout):
        '''
        Function that actually run the test
        '''
        needs = training_matrix(param, plot=False)
        del needs
        self.assertEqual(mock_stdout.getvalue(), exp)

    def test_training_matrix_nohighlevel(self):
        '''
        Test that the training matrix data are correct
        '''

        ###get the needs dictionary
        printout ='\033[1m'+\
                  '[Neo: Matrix Generation:]\033[0m : High Level matrix not found...exit..\n'

        ####test
        self.displayfinal_nohighlevel(self.Fake_matrix2, printout)

    def test_submatrix_topics(self):
        '''
        Test that the submatrix skills are correct
        '''
        ###get the needs dictionary
        skills_prof, skills_Npeople = single_matrix(self.Fake_matrix, 'HL1', 2, 'A', plot=False)

        ###check that the skills are correct
        skills = list(skills_prof.keys())
        skills2 = list(skills_Npeople.keys())


        ##check
        self.assertEqual(skills[0], 'HL1_1')
        self.assertEqual(skills[1], 'HL1_2')
        self.assertEqual(skills[2], 'HL1_3')
        self.assertEqual(skills[3], 'HL1_4')
        self.assertEqual(skills[4], 'HL1_5')
        self.assertEqual(skills[5], 'HL1_6')
        self.assertEqual(skills[6], 'HL1_7')
        self.assertEqual(skills[7], 'HL1_8')
        self.assertEqual(skills[8], 'HL1_9')

        self.assertEqual(skills2[0], 'HL1_1')
        self.assertEqual(skills2[1], 'HL1_2')
        self.assertEqual(skills2[2], 'HL1_3')
        self.assertEqual(skills2[3], 'HL1_4')
        self.assertEqual(skills2[4], 'HL1_5')
        self.assertEqual(skills2[5], 'HL1_6')
        self.assertEqual(skills2[6], 'HL1_7')
        self.assertEqual(skills2[7], 'HL1_8')
        self.assertEqual(skills2[8], 'HL1_9')


    def test_submatrix_data(self):
        '''
        Test that the submatrix skills have the correct data
        '''
        ###get the needs dictionary
        skills_prof, skills_Npeople = single_matrix(self.Fake_matrix, 'HL1', 2, 'A', plot=False)

        ##check
        self.assertEqual(skills_prof['HL1_1'], 14)
        self.assertEqual(skills_prof['HL1_2'], 14)
        self.assertEqual(skills_prof['HL1_3'], 11)
        self.assertEqual(skills_prof['HL1_4'], 9)
        self.assertEqual(skills_prof['HL1_5'], 3)
        self.assertEqual(skills_prof['HL1_6'], 0)
        self.assertEqual(skills_prof['HL1_7'], 0)
        self.assertEqual(skills_prof['HL1_8'], 8)
        self.assertEqual(skills_prof['HL1_9'], 10)



        ##check
        self.assertEqual(skills_Npeople['HL1_1'], 5)
        self.assertEqual(skills_Npeople['HL1_2'], 5)
        self.assertEqual(skills_Npeople['HL1_3'], 5)
        self.assertEqual(skills_Npeople['HL1_4'], 4)
        self.assertEqual(skills_Npeople['HL1_5'], 3)
        self.assertEqual(skills_Npeople['HL1_6'], 0)
        self.assertEqual(skills_Npeople['HL1_7'], 0)
        self.assertEqual(skills_Npeople['HL1_8'], 5)
        self.assertEqual(skills_Npeople['HL1_9'], 5)


    def test_submatrix_data_wrong_column(self):
        '''
        Test when one of the column is wrongly filled (missing data)
        '''
        ###get the needs dictionary
        skills_prof, skills_Npeople = single_matrix(self.Fake_matrix, 'HL4', 2, 'A', plot=False)

        ##check
        self.assertEqual(skills_prof['HL4_1'], 6)
        self.assertEqual(skills_prof['HL4_2'], 0)
        self.assertEqual(skills_prof['HL4_3'], 0)
        self.assertEqual(skills_prof['HL4_4'], 1)
        self.assertEqual(skills_prof['HL4_5'], 0)

        ##check
        self.assertEqual(skills_Npeople['HL4_1'], 2)
        self.assertEqual(skills_Npeople['HL4_2'], 0)
        self.assertEqual(skills_Npeople['HL4_3'], 0)
        self.assertEqual(skills_Npeople['HL4_4'], 1)
        self.assertEqual(skills_Npeople['HL4_5'], 0)


    @unittest.mock.patch('sys.stdout', new_callable=io.StringIO)
    def displayfinal_wrongsubmatrixname(self, param, exp, mock_stdout):
        '''
        Function that actually run the test and catch the output
        '''
        skills_prof, skills_Npeople = single_matrix(param, 'HL5', 2, 'A', plot=False)
        del skills_prof, skills_Npeople
        self.assertEqual(mock_stdout.getvalue(), exp)

    def test_training_matrix_wrongsubmatrixname(self):
        '''
        start the test
        '''
        ###get the needs dictionary
        printout ='\033[1m'+'[Neo: Matrix Generation:]\033[0m : Matrix HL5 not found...exit...\n'

        ####test
        self.displayfinal_wrongsubmatrixname(self.Fake_matrix2, printout)




    def test_rsematrix_topics(self):
        '''
        Test that the single rse matrix skills are correct
        '''
        ###get the needs dictionary
        rse_skills = single_rse(self.Fake_matrix, 'HL1', 2, 'A', 'Rogers', plot=False)

        ###check that the skills are correct
        skills = list(rse_skills.keys())


        ##check
        self.assertEqual(skills[0], 'HL1_1')
        self.assertEqual(skills[1], 'HL1_2')
        self.assertEqual(skills[2], 'HL1_3')
        self.assertEqual(skills[3], 'HL1_4')
        self.assertEqual(skills[4], 'HL1_5')
        self.assertEqual(skills[5], 'HL1_6')
        self.assertEqual(skills[6], 'HL1_7')
        self.assertEqual(skills[7], 'HL1_8')
        self.assertEqual(skills[8], 'HL1_9')

    def test_rsematrix_data(self):
        '''
        Test that the single rse matrix have the correct data
        '''
        ###get the needs dictionary
        rse_skills = single_rse(self.Fake_matrix, 'HL1', 2, 'A', 'Rogers', plot=False)

        ##check Profixienxy
        self.assertEqual(rse_skills['HL1_1']['Proficiency'], 3)
        self.assertEqual(rse_skills['HL1_2']['Proficiency'], 3)
        self.assertEqual(rse_skills['HL1_3']['Proficiency'], 3)
        self.assertEqual(rse_skills['HL1_4']['Proficiency'], 2)
        self.assertEqual(rse_skills['HL1_5']['Proficiency'], 1)
        self.assertEqual(rse_skills['HL1_6']['Proficiency'], 0)
        self.assertEqual(rse_skills['HL1_7']['Proficiency'], 0)
        self.assertEqual(rse_skills['HL1_8']['Proficiency'], 1)
        self.assertEqual(rse_skills['HL1_9']['Proficiency'], 2)

        ##check usage
        self.assertEqual(rse_skills['HL1_1']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_2']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_3']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_4']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_5']['usage'], 'No')
        self.assertEqual(rse_skills['HL1_6']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_7']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_8']['usage'], 'Yes')
        self.assertEqual(rse_skills['HL1_9']['usage'], 'No')



    @unittest.mock.patch('sys.stdout', new_callable=io.StringIO)
    def displayfinal_wrongrsename(self, param, exp, mock_stdout):
        '''
        Function that actually run the test and catch the output
        '''
        rse_skills = single_rse(param, 'HL1', 2, 'A', 'romain', plot=False)
        del rse_skills
        self.assertEqual(mock_stdout.getvalue(), exp)

    def test_training_matrix_wrongrsename(self):
        '''
        start the test
        '''
        ###get the needs dictionary
        printout ='\033[1m'+\
                  '[Neo: Matrix Generation:]\033[0m : The HL1 matrix for romain is empty...exit...\n'

        ####test
        self.displayfinal_wrongrsename(self.Fake_matrix, printout)
